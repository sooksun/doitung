"""Export DE workshop teacher login credentials to an XLSX workbook.

Reads scripts/de-list.json (the original list of 127 attendees) and writes a
workbook with one sheet per school plus an Index sheet. Email pattern follows
production: teacher{N}@{school_code}.local; password is the seeder default
"Teacher@123".

Run:
    py scripts/export-de-credentials.py
Output:
    scripts/DE-teacher-credentials.xlsx
"""
from __future__ import annotations

import json
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
SRC = ROOT / "de-list.json"
OUT = ROOT / "DE-teacher-credentials.xlsx"

DEFAULT_PASSWORD = "Teacher@123"

# Excel school-name -> (sheet tab name, prod school code).
# The prod codes were confirmed via docker exec on the production server.
SCHOOL_MAP = OrderedDict([
    ("โรงเรียนบ้านแม่หม้อ",                 ("บ้านแม่หม้อ",              "57030130")),
    ("โรงเรียนบ้านพญาไพร",                  ("บ้านพญาไพร",              "57030129")),
    ("โรงเรียนสามัคคีพัฒนา",                ("สามัคคีพัฒนา",             "57030134")),
    ("โรงเรียนตำรวจตระเวนชายแดนบำรุงที่ 87", ("ตชด.บำรุงที่ 87",          "57030181")),
    ("โรงเรียนบ้านปางมะหัน",                ("บ้านปางมะหัน",            "57030135")),
    ("โรงเรียนบ้านห้วยอิ้น",                ("บ้านห้วยอื้น",             "57030136")),
    ("โรงเรียนบ้านผาจี",                    ("บ้านผาจี",                "57030139")),
    ("โรงเรียนบ้านห้วยหยวกป่าโซ",          ("บ้านห้วยหยวกป่าโซ",        "57030175")),
    ("โรงเรียนบ้านห้วยไร่สามัคคี",         ("บ้านห้วยไร่สามัคคี",      "57030143")),
])

# --- styles -------------------------------------------------------------------

THIN = Side(border_style="thin", color="CFD8DC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TITLE_FONT = Font(name="TH Sarabun New", size=18, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", start_color="4F46E5")  # indigo
HDR_FONT = Font(name="TH Sarabun New", size=14, bold=True, color="1F2937")
HDR_FILL = PatternFill("solid", start_color="E0E7FF")
ROW_FONT = Font(name="TH Sarabun New", size=14)
MONO_FONT = Font(name="Consolas", size=12)


def style_sheet(ws, title: str, total: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    sub = ws.cell(row=2, column=1, value=f"รวม {total} คน · รหัสผ่านเริ่มต้น: {DEFAULT_PASSWORD}")
    sub.font = Font(name="TH Sarabun New", size=12, italic=True, color="6B7280")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    ws.row_dimensions[2].height = 20

    headers = ["ลำดับ", "ชื่อ-นามสกุล", "Username", "Password"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[3].height = 24

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 16

    ws.freeze_panes = "A4"


def write_rows(ws, teachers: list[dict], code: str) -> None:
    for i, t in enumerate(teachers, start=1):
        username = f"teacher{i}@{code}.local"
        row = 3 + i
        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=1).font = ROW_FONT
        ws.cell(row=row, column=2, value=t["ชื่อ-นามสกุล"]).font = ROW_FONT
        ws.cell(row=row, column=3, value=username).font = MONO_FONT
        ws.cell(row=row, column=4, value=DEFAULT_PASSWORD).font = MONO_FONT
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = BORDER
            ws.cell(row=row, column=col).alignment = Alignment(vertical="center")
        ws.cell(row=row, column=2).alignment = Alignment(vertical="center", indent=1)


# --- main ---------------------------------------------------------------------

def main() -> None:
    data = json.loads(SRC.read_text(encoding="utf-8"))
    rows = data["sheets"][0]["rows"]
    rows = [r for r in rows if r.get("โรงเรียน") and r.get("ชื่อ-นามสกุล")]

    grouped: dict[str, list[dict]] = OrderedDict()
    for r in rows:
        key = r["โรงเรียน"]
        grouped.setdefault(key, []).append(r)

    wb = Workbook()
    # default sheet -> use as Index
    idx = wb.active
    idx.title = "สรุป"

    # Header
    idx.merge_cells("A1:D1")
    c = idx.cell(row=1, column=1, value="รายชื่อผู้เข้าร่วมประชุม DE — ชื่อผู้ใช้และรหัสผ่าน")
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    idx.row_dimensions[1].height = 36

    sub = idx.cell(row=2, column=1, value=f"รหัสผ่านเริ่มต้นทุกบัญชี: {DEFAULT_PASSWORD}  ·  URL: https://doitung.cnppai.com/login")
    sub.font = Font(name="TH Sarabun New", size=13, italic=True, color="6B7280")
    idx.merge_cells("A2:D2")
    idx.row_dimensions[2].height = 22

    for i, h in enumerate(["ลำดับ", "โรงเรียน", "จำนวนครู", "Email pattern"], start=1):
        c = idx.cell(row=3, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    idx.row_dimensions[3].height = 24

    idx.column_dimensions["A"].width = 8
    idx.column_dimensions["B"].width = 40
    idx.column_dimensions["C"].width = 12
    idx.column_dimensions["D"].width = 36

    total_teachers = 0
    line = 4
    seen_schools = 0

    for excel_name, mapped in SCHOOL_MAP.items():
        sheet_name, code = mapped
        teachers = grouped.get(excel_name, [])
        if not teachers:
            continue
        seen_schools += 1
        total_teachers += len(teachers)

        # Add to index
        idx.cell(row=line, column=1, value=seen_schools).alignment = Alignment(horizontal="center")
        idx.cell(row=line, column=1).font = ROW_FONT
        idx.cell(row=line, column=2, value=excel_name).font = ROW_FONT
        idx.cell(row=line, column=3, value=f"{len(teachers)} คน").alignment = Alignment(horizontal="center")
        idx.cell(row=line, column=3).font = ROW_FONT
        idx.cell(row=line, column=4, value=f"teacher1..{len(teachers)}@{code}.local").font = MONO_FONT
        for col in range(1, 5):
            idx.cell(row=line, column=col).border = BORDER
        line += 1

        ws = wb.create_sheet(title=sheet_name[:31])
        style_sheet(ws, excel_name, len(teachers))
        write_rows(ws, teachers, code)

    # Total row
    idx.cell(row=line, column=2, value="รวมทั้งหมด").font = Font(name="TH Sarabun New", size=14, bold=True)
    idx.cell(row=line, column=2).alignment = Alignment(horizontal="right")
    idx.cell(row=line, column=3, value=f"{total_teachers} คน").font = Font(name="TH Sarabun New", size=14, bold=True)
    idx.cell(row=line, column=3).alignment = Alignment(horizontal="center")
    idx.cell(row=line, column=3).fill = HDR_FILL

    idx.freeze_panes = "A4"

    wb.save(OUT)
    print(f"Wrote {OUT} — {seen_schools} schools, {total_teachers} teachers")


if __name__ == "__main__":
    main()
